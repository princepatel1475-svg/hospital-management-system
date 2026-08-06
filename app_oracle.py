"""
Hospital Management System (Oracle Edition)
--------------------------------------------
A GUI-based hospital management system with Oracle DB backend and
live Excel synchronization.

Requirements: oracledb, openpyxl
Install with: pip install oracledb openpyxl

Uses python-oracledb in "thin" mode -- no Oracle Instant Client needed.
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import date, datetime
import oracledb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# ============================================
# DATABASE CONFIGURATION - EDIT THESE VALUES
# ============================================
import os
DB_CONFIG = {
    "user": "system",
    "password": os.environ.get("ORACLE_DB_PASSWORD", ""),
    "dsn": "localhost:1521/XE"
}

EXCEL_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hospital_data_export.xlsx")


class Database:
    """Handles all Oracle database operations."""

    def __init__(self, config):
        self.config = config
        self.conn = None
        self.connect()

    def connect(self):
        try:
            self.conn = oracledb.connect(
                user=self.config["user"],
                password=self.config["password"],
                dsn=self.config["dsn"]
            )
        except oracledb.Error as e:
            messagebox.showerror("Database Connection Error", str(e))
            raise

    def query(self, sql, params=None):
        """Run a SELECT and return list of dicts."""
        cur = self.conn.cursor()
        cur.execute(sql, params or {})
        cols = [c[0].lower() for c in cur.description]
        rows = [dict(zip(cols, row)) for row in cur.fetchall()]
        cur.close()
        return rows

    def execute(self, sql, params=None, commit=False):
        """Run INSERT/UPDATE/DELETE."""
        cur = self.conn.cursor()
        cur.execute(sql, params or {})
        if commit:
            self.conn.commit()
        cur.close()

    def close(self):
        if self.conn:
            self.conn.close()


class HospitalApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Hospital Management System (Oracle)")
        self.geometry("1000x650")
        self.configure(bg="#f4f6f8")

        try:
            self.db = Database(DB_CONFIG)
        except Exception:
            self.destroy()
            return

        self.create_widgets()
        self.refresh_all()

    # ------------------------------------------------------------
    # UI SETUP
    # ------------------------------------------------------------
    def create_widgets(self):
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview", rowheight=26, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))

        header = tk.Frame(self, bg="#b23c17", height=60)
        header.pack(fill="x")
        tk.Label(header, text="🏥 Hospital Management System", bg="#b23c17",
                 fg="white", font=("Segoe UI", 18, "bold")).pack(pady=12, padx=15, anchor="w")

        self.tabs = ttk.Notebook(self)
        self.tabs.pack(fill="both", expand=True, padx=10, pady=10)

        self.patients_tab = ttk.Frame(self.tabs)
        self.staff_tab = ttk.Frame(self.tabs)
        self.doctors_tab = ttk.Frame(self.tabs)
        self.appointments_tab = ttk.Frame(self.tabs)
        self.billing_tab = ttk.Frame(self.tabs)
        self.pharmacy_tab = ttk.Frame(self.tabs)
        self.reports_tab = ttk.Frame(self.tabs)

        self.tabs.add(self.patients_tab, text="Patients")
        self.tabs.add(self.staff_tab, text="Staff")
        self.tabs.add(self.doctors_tab, text="Doctors")
        self.tabs.add(self.appointments_tab, text="Appointments")
        self.tabs.add(self.billing_tab, text="Billing")
        self.tabs.add(self.pharmacy_tab, text="Pharmacy / Medical Store")
        self.tabs.add(self.reports_tab, text="Reports / Excel Export")

        self.build_patients_tab()
        self.build_staff_tab()
        self.build_doctors_tab()
        self.build_appointments_tab()
        self.build_billing_tab()
        self.build_pharmacy_tab()
        self.build_reports_tab()

    # ------------------------------------------------------------
    # PATIENTS TAB
    # ------------------------------------------------------------
    def build_patients_tab(self):
        form = tk.LabelFrame(self.patients_tab, text="Add / Update Patient", padx=10, pady=10)
        form.pack(fill="x", padx=10, pady=10)

        labels = ["First Name", "Last Name", "DOB (YYYY-MM-DD)", "Gender",
                  "Phone", "Email", "Address", "Blood Group"]
        self.patient_entries = {}
        for i, lbl in enumerate(labels):
            r, c = divmod(i, 4)
            tk.Label(form, text=lbl).grid(row=r*2, column=c, sticky="w", padx=5)
            if lbl == "Gender":
                var = ttk.Combobox(form, values=["Male", "Female", "Other"], width=18, state="readonly")
            else:
                var = tk.Entry(form, width=20)
            var.grid(row=r*2+1, column=c, padx=5, pady=3)
            self.patient_entries[lbl] = var

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=4, column=0, columnspan=4, pady=8, sticky="w")
        tk.Button(btn_frame, text="Add Patient", bg="#2e7d32", fg="white",
                  command=self.add_patient).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Delete Selected", bg="#c62828", fg="white",
                  command=self.delete_patient).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Clear Form", command=lambda: self.clear_entries(self.patient_entries)
                  ).pack(side="left", padx=5)

        cols = ("ID", "First", "Last", "DOB", "Gender", "Phone", "Email", "Address", "Blood")
        self.patients_tree = ttk.Treeview(self.patients_tab, columns=cols, show="headings", height=14)
        for c in cols:
            self.patients_tree.heading(c, text=c)
            self.patients_tree.column(c, width=100, anchor="w")
        self.patients_tree.pack(fill="both", expand=True, padx=10, pady=5)

    def add_patient(self):
        e = self.patient_entries
        try:
            self.db.execute(
                """INSERT INTO patients (first_name, last_name, date_of_birth, gender,
                   phone, email, address, blood_group)
                   VALUES (:1, :2, TO_DATE(:3,'YYYY-MM-DD'), :4, :5, :6, :7, :8)""",
                [e["First Name"].get(), e["Last Name"].get(), e["DOB (YYYY-MM-DD)"].get(),
                 e["Gender"].get(), e["Phone"].get(), e["Email"].get(),
                 e["Address"].get(), e["Blood Group"].get()],
                commit=True
            )
            messagebox.showinfo("Success", "Patient added.")
            self.clear_entries(e)
            self.refresh_all()
        except oracledb.Error as err:
            messagebox.showerror("Error", str(err))

    def delete_patient(self):
        sel = self.patients_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a patient row first.")
            return
        patient_id = self.patients_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", f"Delete patient ID {patient_id}? This also removes their appointments."):
            self.db.execute("DELETE FROM patients WHERE patient_id=:1", [patient_id], commit=True)
            self.refresh_all()

    # ------------------------------------------------------------
    # STAFF TAB
    # ------------------------------------------------------------
    def build_staff_tab(self):
        form = tk.LabelFrame(self.staff_tab, text="Add / Update Staff", padx=10, pady=10)
        form.pack(fill="x", padx=10, pady=10)

        labels = ["First Name", "Last Name", "Role", "Department", "Phone", "Email", "Hire Date (YYYY-MM-DD)"]
        self.staff_entries = {}
        for i, lbl in enumerate(labels):
            r, c = divmod(i, 4)
            tk.Label(form, text=lbl).grid(row=r*2, column=c, sticky="w", padx=5)
            if lbl == "Role":
                var = ttk.Combobox(form, values=["Doctor", "Nurse", "Receptionist", "Admin", "Technician"],
                                    width=18, state="readonly")
            else:
                var = tk.Entry(form, width=20)
            var.grid(row=r*2+1, column=c, padx=5, pady=3)
            self.staff_entries[lbl] = var

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=4, column=0, columnspan=4, pady=8, sticky="w")
        tk.Button(btn_frame, text="Add Staff", bg="#2e7d32", fg="white",
                  command=self.add_staff).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Delete Selected", bg="#c62828", fg="white",
                  command=self.delete_staff).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Clear Form", command=lambda: self.clear_entries(self.staff_entries)
                  ).pack(side="left", padx=5)

        cols = ("ID", "First", "Last", "Role", "Department", "Phone", "Email", "Hire Date")
        self.staff_tree = ttk.Treeview(self.staff_tab, columns=cols, show="headings", height=14)
        for c in cols:
            self.staff_tree.heading(c, text=c)
            self.staff_tree.column(c, width=100, anchor="w")
        self.staff_tree.pack(fill="both", expand=True, padx=10, pady=5)

    def add_staff(self):
        e = self.staff_entries
        try:
            self.db.execute(
                """INSERT INTO staff (first_name, last_name, role, department, phone, email, hire_date)
                   VALUES (:1, :2, :3, :4, :5, :6, TO_DATE(:7,'YYYY-MM-DD'))""",
                [e["First Name"].get(), e["Last Name"].get(), e["Role"].get(),
                 e["Department"].get(), e["Phone"].get(), e["Email"].get(),
                 e["Hire Date (YYYY-MM-DD)"].get()],
                commit=True
            )
            messagebox.showinfo("Success", "Staff member added.")
            self.clear_entries(e)
            self.refresh_all()
        except oracledb.Error as err:
            messagebox.showerror("Error", str(err))

    def delete_staff(self):
        sel = self.staff_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a staff row first.")
            return
        staff_id = self.staff_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", f"Delete staff ID {staff_id}?"):
            try:
                self.db.execute("DELETE FROM staff WHERE staff_id=:1", [staff_id], commit=True)
                self.refresh_all()
            except oracledb.Error as err:
                messagebox.showerror("Error", f"Cannot delete: {err}\n(They may have existing appointments.)")

    # ------------------------------------------------------------
    # DOCTORS TAB
    # ------------------------------------------------------------
    def build_doctors_tab(self):
        form = tk.LabelFrame(self.doctors_tab, text="Add / Update Doctor", padx=10, pady=10)
        form.pack(fill="x", padx=10, pady=10)

        labels = ["First Name", "Last Name", "Specialization", "Qualification",
                  "Phone", "Email", "Consultation Fee", "Availability Days"]
        self.doctor_entries = {}
        for i, lbl in enumerate(labels):
            r, c = divmod(i, 4)
            tk.Label(form, text=lbl).grid(row=r*2, column=c, sticky="w", padx=5)
            var = tk.Entry(form, width=20)
            var.grid(row=r*2+1, column=c, padx=5, pady=3)
            self.doctor_entries[lbl] = var

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=4, column=0, columnspan=4, pady=8, sticky="w")
        tk.Button(btn_frame, text="Add Doctor", bg="#2e7d32", fg="white",
                  command=self.add_doctor).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Delete Selected", bg="#c62828", fg="white",
                  command=self.delete_doctor).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Clear Form", command=lambda: self.clear_entries(self.doctor_entries)
                  ).pack(side="left", padx=5)

        cols = ("ID", "First", "Last", "Specialization", "Qualification", "Phone", "Email", "Fee", "Availability")
        self.doctors_tree = ttk.Treeview(self.doctors_tab, columns=cols, show="headings", height=14)
        for c in cols:
            self.doctors_tree.heading(c, text=c)
            self.doctors_tree.column(c, width=100, anchor="w")
        self.doctors_tree.pack(fill="both", expand=True, padx=10, pady=5)

    def add_doctor(self):
        e = self.doctor_entries
        try:
            self.db.execute(
                """INSERT INTO doctors (first_name, last_name, specialization, qualification,
                   phone, email, consultation_fee, availability_days)
                   VALUES (:1, :2, :3, :4, :5, :6, :7, :8)""",
                [e["First Name"].get(), e["Last Name"].get(), e["Specialization"].get(),
                 e["Qualification"].get(), e["Phone"].get(), e["Email"].get(),
                 e["Consultation Fee"].get(), e["Availability Days"].get()],
                commit=True
            )
            messagebox.showinfo("Success", "Doctor added.")
            self.clear_entries(e)
            self.refresh_all()
        except oracledb.Error as err:
            messagebox.showerror("Error", str(err))

    def delete_doctor(self):
        sel = self.doctors_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a doctor row first.")
            return
        doctor_id = self.doctors_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", f"Delete doctor ID {doctor_id}?"):
            self.db.execute("DELETE FROM doctors WHERE doctor_id=:1", [doctor_id], commit=True)
            self.refresh_all()

    # ------------------------------------------------------------
    # APPOINTMENTS TAB
    # ------------------------------------------------------------
    def build_appointments_tab(self):
        form = tk.LabelFrame(self.appointments_tab, text="Add / Update Appointment", padx=10, pady=10)
        form.pack(fill="x", padx=10, pady=10)

        labels = ["Patient ID", "Staff ID", "Date (YYYY-MM-DD)", "Time (HH:MM:SS)",
                  "Reason", "Status"]
        self.appt_entries = {}
        for i, lbl in enumerate(labels):
            r, c = divmod(i, 3)
            tk.Label(form, text=lbl).grid(row=r*2, column=c, sticky="w", padx=5)
            if lbl == "Status":
                var = ttk.Combobox(form, values=["Scheduled", "Completed", "Cancelled", "No-Show"],
                                    width=20, state="readonly")
                var.set("Scheduled")
            else:
                var = tk.Entry(form, width=22)
            var.grid(row=r*2+1, column=c, padx=5, pady=3)
            self.appt_entries[lbl] = var

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=8, sticky="w")
        tk.Button(btn_frame, text="Add Appointment", bg="#2e7d32", fg="white",
                  command=self.add_appointment).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Delete Selected", bg="#c62828", fg="white",
                  command=self.delete_appointment).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Clear Form", command=lambda: self.clear_entries(self.appt_entries)
                  ).pack(side="left", padx=5)

        cols = ("ID", "Patient", "Staff", "Date", "Time", "Reason", "Status")
        self.appt_tree = ttk.Treeview(self.appointments_tab, columns=cols, show="headings", height=14)
        for c in cols:
            self.appt_tree.heading(c, text=c)
            self.appt_tree.column(c, width=110, anchor="w")
        self.appt_tree.pack(fill="both", expand=True, padx=10, pady=5)

    def add_appointment(self):
        e = self.appt_entries
        try:
            self.db.execute(
                """INSERT INTO appointments (patient_id, staff_id, appointment_date,
                   appointment_time, reason, status)
                   VALUES (:1, :2, TO_DATE(:3,'YYYY-MM-DD'), :4, :5, :6)""",
                [e["Patient ID"].get(), e["Staff ID"].get(), e["Date (YYYY-MM-DD)"].get(),
                 e["Time (HH:MM:SS)"].get(), e["Reason"].get(), e["Status"].get()],
                commit=True
            )
            messagebox.showinfo("Success", "Appointment added.")
            self.clear_entries(e)
            self.refresh_all()
        except oracledb.Error as err:
            messagebox.showerror("Error", str(err))

    def delete_appointment(self):
        sel = self.appt_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select an appointment row first.")
            return
        appt_id = self.appt_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", f"Delete appointment ID {appt_id}?"):
            self.db.execute("DELETE FROM appointments WHERE appointment_id=:1", [appt_id], commit=True)
            self.refresh_all()

    # ------------------------------------------------------------
    # BILLING TAB
    # ------------------------------------------------------------
    def build_billing_tab(self):
        form = tk.LabelFrame(self.billing_tab, text="Create Bill", padx=10, pady=10)
        form.pack(fill="x", padx=10, pady=10)

        labels = ["Patient ID", "Appointment ID (optional)", "Consultation Charge",
                  "Medicine Charge", "Room Charge", "Other Charge", "Payment Status", "Payment Method"]
        self.bill_entries = {}
        for i, lbl in enumerate(labels):
            r, c = divmod(i, 4)
            tk.Label(form, text=lbl).grid(row=r*2, column=c, sticky="w", padx=5)
            if lbl == "Payment Status":
                var = ttk.Combobox(form, values=["Pending", "Paid", "Partial"], width=18, state="readonly")
                var.set("Pending")
            elif lbl == "Payment Method":
                var = ttk.Combobox(form, values=["Cash", "Card", "UPI", "Insurance"], width=18, state="readonly")
            else:
                var = tk.Entry(form, width=20)
            var.grid(row=r*2+1, column=c, padx=5, pady=3)
            self.bill_entries[lbl] = var

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=4, column=0, columnspan=4, pady=8, sticky="w")
        tk.Button(btn_frame, text="Create Bill", bg="#2e7d32", fg="white",
                  command=self.add_bill).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Delete Selected", bg="#c62828", fg="white",
                  command=self.delete_bill).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Clear Form", command=lambda: self.clear_entries(self.bill_entries)
                  ).pack(side="left", padx=5)

        cols = ("ID", "Patient ID", "Appt ID", "Date", "Consult", "Medicine", "Room", "Other", "Total", "Status", "Method")
        self.bills_tree = ttk.Treeview(self.billing_tab, columns=cols, show="headings", height=13)
        for c in cols:
            self.bills_tree.heading(c, text=c)
            self.bills_tree.column(c, width=85, anchor="w")
        self.bills_tree.pack(fill="both", expand=True, padx=10, pady=5)

    def add_bill(self):
        e = self.bill_entries
        appt_id = e["Appointment ID (optional)"].get().strip() or None
        try:
            self.db.execute(
                """INSERT INTO bills (patient_id, appointment_id, consultation_charge,
                   medicine_charge, room_charge, other_charge, payment_status, payment_method)
                   VALUES (:1, :2, :3, :4, :5, :6, :7, :8)""",
                [e["Patient ID"].get(), appt_id,
                 e["Consultation Charge"].get() or 0, e["Medicine Charge"].get() or 0,
                 e["Room Charge"].get() or 0, e["Other Charge"].get() or 0,
                 e["Payment Status"].get(), e["Payment Method"].get() or None],
                commit=True
            )
            messagebox.showinfo("Success", "Bill created.")
            self.clear_entries(e)
            self.refresh_all()
        except oracledb.Error as err:
            messagebox.showerror("Error", str(err))

    def delete_bill(self):
        sel = self.bills_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a bill row first.")
            return
        bill_id = self.bills_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", f"Delete bill ID {bill_id}?"):
            self.db.execute("DELETE FROM bills WHERE bill_id=:1", [bill_id], commit=True)
            self.refresh_all()

    # ------------------------------------------------------------
    # PHARMACY / MEDICAL STORE TAB
    # ------------------------------------------------------------
    def build_pharmacy_tab(self):
        inv_form = tk.LabelFrame(self.pharmacy_tab, text="Add / Update Medicine (Inventory)", padx=10, pady=10)
        inv_form.pack(fill="x", padx=10, pady=8)

        labels = ["Medicine Name", "Category", "Unit Price", "Stock Quantity", "Expiry Date (YYYY-MM-DD)"]
        self.medicine_entries = {}
        for i, lbl in enumerate(labels):
            tk.Label(inv_form, text=lbl).grid(row=0, column=i, sticky="w", padx=5)
            var = tk.Entry(inv_form, width=18)
            var.grid(row=1, column=i, padx=5, pady=3)
            self.medicine_entries[lbl] = var
        btn1 = tk.Frame(inv_form)
        btn1.grid(row=2, column=0, columnspan=5, pady=6, sticky="w")
        tk.Button(btn1, text="Add Medicine", bg="#2e7d32", fg="white",
                  command=self.add_medicine).pack(side="left", padx=5)
        tk.Button(btn1, text="Delete Selected", bg="#c62828", fg="white",
                  command=self.delete_medicine).pack(side="left", padx=5)

        cols1 = ("ID", "Name", "Category", "Unit Price", "Stock", "Expiry")
        self.medicines_tree = ttk.Treeview(self.pharmacy_tab, columns=cols1, show="headings", height=6)
        for c in cols1:
            self.medicines_tree.heading(c, text=c)
            self.medicines_tree.column(c, width=110, anchor="w")
        self.medicines_tree.pack(fill="x", padx=10, pady=5)

        sale_form = tk.LabelFrame(self.pharmacy_tab, text="Record Sale (Medical Store Revenue)", padx=10, pady=10)
        sale_form.pack(fill="x", padx=10, pady=8)
        labels2 = ["Medicine ID", "Patient ID (optional)", "Quantity", "Total Amount"]
        self.sale_entries = {}
        for i, lbl in enumerate(labels2):
            tk.Label(sale_form, text=lbl).grid(row=0, column=i, sticky="w", padx=5)
            var = tk.Entry(sale_form, width=18)
            var.grid(row=1, column=i, padx=5, pady=3)
            self.sale_entries[lbl] = var
        btn2 = tk.Frame(sale_form)
        btn2.grid(row=2, column=0, columnspan=4, pady=6, sticky="w")
        tk.Button(btn2, text="Record Sale", bg="#2e7d32", fg="white",
                  command=self.add_sale).pack(side="left", padx=5)

        cols2 = ("Sale ID", "Medicine ID", "Patient ID", "Qty", "Date", "Amount")
        self.sales_tree = ttk.Treeview(self.pharmacy_tab, columns=cols2, show="headings", height=6)
        for c in cols2:
            self.sales_tree.heading(c, text=c)
            self.sales_tree.column(c, width=110, anchor="w")
        self.sales_tree.pack(fill="x", padx=10, pady=5)

        self.pharmacy_revenue_lbl = tk.Label(self.pharmacy_tab, text="Total Pharmacy Revenue: 0",
                                              font=("Segoe UI", 12, "bold"), fg="#b23c17")
        self.pharmacy_revenue_lbl.pack(pady=8, anchor="w", padx=10)

    def add_medicine(self):
        e = self.medicine_entries
        try:
            self.db.execute(
                """INSERT INTO medicines (medicine_name, category, unit_price, stock_quantity, expiry_date)
                   VALUES (:1, :2, :3, :4, TO_DATE(:5,'YYYY-MM-DD'))""",
                [e["Medicine Name"].get(), e["Category"].get(), e["Unit Price"].get(),
                 e["Stock Quantity"].get() or 0, e["Expiry Date (YYYY-MM-DD)"].get()],
                commit=True
            )
            messagebox.showinfo("Success", "Medicine added.")
            self.clear_entries(e)
            self.refresh_all()
        except oracledb.Error as err:
            messagebox.showerror("Error", str(err))

    def delete_medicine(self):
        sel = self.medicines_tree.selection()
        if not sel:
            messagebox.showwarning("No selection", "Select a medicine row first.")
            return
        med_id = self.medicines_tree.item(sel[0])["values"][0]
        if messagebox.askyesno("Confirm", f"Delete medicine ID {med_id}?"):
            self.db.execute("DELETE FROM medicines WHERE medicine_id=:1", [med_id], commit=True)
            self.refresh_all()

    def add_sale(self):
        e = self.sale_entries
        patient_id = e["Patient ID (optional)"].get().strip() or None
        try:
            self.db.execute(
                """INSERT INTO medicine_sales (medicine_id, patient_id, quantity, total_amount)
                   VALUES (:1, :2, :3, :4)""",
                [e["Medicine ID"].get(), patient_id, e["Quantity"].get(), e["Total Amount"].get()],
                commit=True
            )
            messagebox.showinfo("Success", "Sale recorded and stock updated.")
            self.clear_entries(e)
            self.refresh_all()
        except oracledb.Error as err:
            messagebox.showerror("Error", str(err))

    # ------------------------------------------------------------
    # REPORTS / EXCEL TAB
    # ------------------------------------------------------------
    def build_reports_tab(self):
        info = tk.Label(self.reports_tab,
                         text="Every change made through this GUI automatically updates the linked Excel file.\n"
                              "Click below any time to force a fresh export / download copy.",
                         justify="left", font=("Segoe UI", 10))
        info.pack(padx=15, pady=15, anchor="w")

        tk.Button(self.reports_tab, text="⬇ Export / Refresh Excel Now", bg="#b23c17", fg="white",
                  font=("Segoe UI", 11, "bold"), command=self.export_to_excel).pack(pady=10)

        self.stats_frame = tk.LabelFrame(self.reports_tab, text="Live Summary Statistics", padx=15, pady=15)
        self.stats_frame.pack(fill="x", padx=15, pady=15)

        self.stat_labels = {}
        stat_names = ["Total Patients", "Total Staff", "Total Doctors", "Total Appointments",
                      "Appointments Today", "Scheduled", "Completed", "Cancelled",
                      "Total Billing Revenue", "Pending Bills", "Total Pharmacy Revenue"]
        for i, name in enumerate(stat_names):
            r, c = divmod(i, 4)
            frame = tk.Frame(self.stats_frame, bg="#fdeee7", padx=10, pady=8)
            frame.grid(row=r, column=c, padx=8, pady=8, sticky="nsew")
            tk.Label(frame, text=name, bg="#fdeee7", font=("Segoe UI", 9)).pack()
            val_lbl = tk.Label(frame, text="0", bg="#fdeee7", font=("Segoe UI", 16, "bold"), fg="#b23c17")
            val_lbl.pack()
            self.stat_labels[name] = val_lbl

        path_lbl = tk.Label(self.reports_tab, text=f"Excel file location:\n{EXCEL_FILE_PATH}",
                             fg="#555", font=("Segoe UI", 9))
        path_lbl.pack(pady=10, anchor="w", padx=15)

    def get_stats(self):
        patients = self.db.query("SELECT COUNT(*) AS c FROM patients")[0]["c"]
        staff = self.db.query("SELECT COUNT(*) AS c FROM staff")[0]["c"]
        doctors = self.db.query("SELECT COUNT(*) AS c FROM doctors")[0]["c"]
        total_appts = self.db.query("SELECT COUNT(*) AS c FROM appointments")[0]["c"]
        today_appts = self.db.query(
            "SELECT COUNT(*) AS c FROM appointments WHERE appointment_date = TRUNC(SYSDATE)"
        )[0]["c"]
        scheduled = self.db.query("SELECT COUNT(*) AS c FROM appointments WHERE status='Scheduled'")[0]["c"]
        completed = self.db.query("SELECT COUNT(*) AS c FROM appointments WHERE status='Completed'")[0]["c"]
        cancelled = self.db.query("SELECT COUNT(*) AS c FROM appointments WHERE status='Cancelled'")[0]["c"]
        billing_revenue = self.db.query("SELECT NVL(SUM(total_amount),0) AS c FROM bills")[0]["c"]
        pending_bills = self.db.query("SELECT COUNT(*) AS c FROM bills WHERE payment_status='Pending'")[0]["c"]
        pharmacy_revenue = self.db.query("SELECT NVL(SUM(total_amount),0) AS c FROM medicine_sales")[0]["c"]
        return {
            "Total Patients": patients, "Total Staff": staff, "Total Doctors": doctors,
            "Total Appointments": total_appts,
            "Appointments Today": today_appts, "Scheduled": scheduled,
            "Completed": completed, "Cancelled": cancelled,
            "Total Billing Revenue": billing_revenue, "Pending Bills": pending_bills,
            "Total Pharmacy Revenue": pharmacy_revenue
        }

    def update_stats_display(self):
        stats = self.get_stats()
        for name, val in stats.items():
            self.stat_labels[name].config(text=str(val))

    # ------------------------------------------------------------
    # EXCEL SYNC
    # ------------------------------------------------------------
    def export_to_excel(self):
        self._write_workbook()
        self.update_stats_display()
        messagebox.showinfo("Excel Updated", f"Excel file saved/refreshed at:\n{EXCEL_FILE_PATH}")

    def export_to_excel_silent(self):
        """Same export, no popup -- used for auto-sync after every GUI change."""
        self._write_workbook()

    def _write_workbook(self):
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="B23C17", end_color="B23C17", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws = wb.active
        ws.title = "Patients"
        self._write_sheet(ws, self.db.query("SELECT * FROM patients ORDER BY patient_id"), header_fill, header_font)

        ws2 = wb.create_sheet("Staff")
        self._write_sheet(ws2, self.db.query("SELECT * FROM staff ORDER BY staff_id"), header_fill, header_font)

        ws2b = wb.create_sheet("Doctors")
        self._write_sheet(ws2b, self.db.query("SELECT * FROM doctors ORDER BY doctor_id"), header_fill, header_font)

        ws3 = wb.create_sheet("Appointments")
        appts = self.db.query("""
            SELECT a.appointment_id, p.first_name AS patient_first, p.last_name AS patient_last,
                   s.first_name AS staff_first, s.last_name AS staff_last,
                   a.appointment_date, a.appointment_time, a.reason, a.status
            FROM appointments a
            JOIN patients p ON a.patient_id = p.patient_id
            JOIN staff s ON a.staff_id = s.staff_id
            ORDER BY a.appointment_date DESC, a.appointment_time
        """)
        self._write_sheet(ws3, appts, header_fill, header_font)

        ws3b = wb.create_sheet("Billing")
        bills = self.db.query("""
            SELECT b.bill_id, p.first_name AS patient_first, p.last_name AS patient_last,
                   b.bill_date, b.consultation_charge, b.medicine_charge, b.room_charge,
                   b.other_charge, b.total_amount, b.payment_status, b.payment_method
            FROM bills b
            JOIN patients p ON b.patient_id = p.patient_id
            ORDER BY b.bill_date DESC
        """)
        self._write_sheet(ws3b, bills, header_fill, header_font)

        ws3c = wb.create_sheet("Pharmacy Inventory")
        self._write_sheet(ws3c, self.db.query("SELECT * FROM medicines ORDER BY medicine_id"), header_fill, header_font)

        ws3d = wb.create_sheet("Pharmacy Sales")
        sales = self.db.query("""
            SELECT s.sale_id, m.medicine_name, s.quantity, s.sale_date, s.total_amount
            FROM medicine_sales s
            JOIN medicines m ON s.medicine_id = m.medicine_id
            ORDER BY s.sale_date DESC
        """)
        self._write_sheet(ws3d, sales, header_fill, header_font)

        ws4 = wb.create_sheet("Summary")
        stats = self.get_stats()
        ws4["A1"] = "Hospital Management System — Summary Report"
        ws4["A1"].font = Font(bold=True, size=14, color="B23C17")
        ws4["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row = 4
        for name, val in stats.items():
            ws4[f"A{row}"] = name
            ws4[f"B{row}"] = val
            ws4[f"A{row}"].font = Font(bold=True)
            row += 1
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 15

        wb.save(EXCEL_FILE_PATH)

    def _write_sheet(self, ws, rows, header_fill, header_font):
        if not rows:
            ws["A1"] = "No data"
            return
        headers = list(rows[0].keys())
        for c_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c_idx, value=h.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for r_idx, row in enumerate(rows, start=2):
            for c_idx, h in enumerate(headers, start=1):
                val = row[h]
                if isinstance(val, (datetime, date)):
                    val = str(val)
                ws.cell(row=r_idx, column=c_idx, value=val)
        for c_idx, h in enumerate(headers, start=1):
            col_letter = chr(64 + c_idx) if c_idx <= 26 else "A"
            ws.column_dimensions[col_letter].width = max(14, len(h) + 4)

    # ------------------------------------------------------------
    # REFRESH / UTILITIES
    # ------------------------------------------------------------
    def refresh_all(self):
        self.refresh_patients()
        self.refresh_staff()
        self.refresh_doctors()
        self.refresh_appointments()
        self.refresh_bills()
        self.refresh_medicines()
        self.refresh_sales()
        self.update_stats_display()
        self.export_to_excel_silent()

    def refresh_doctors(self):
        for row in self.doctors_tree.get_children():
            self.doctors_tree.delete(row)
        for d in self.db.query("SELECT * FROM doctors ORDER BY doctor_id"):
            self.doctors_tree.insert("", "end", values=(
                d["doctor_id"], d["first_name"], d["last_name"], d["specialization"],
                d["qualification"], d["phone"], d["email"], d["consultation_fee"], d["availability_days"]
            ))

    def refresh_bills(self):
        for row in self.bills_tree.get_children():
            self.bills_tree.delete(row)
        rows = self.db.query("""
            SELECT b.bill_id, b.patient_id, b.appointment_id, b.bill_date,
                   b.consultation_charge, b.medicine_charge, b.room_charge, b.other_charge,
                   b.total_amount, b.payment_status, b.payment_method
            FROM bills b ORDER BY b.bill_date DESC
        """)
        for b in rows:
            self.bills_tree.insert("", "end", values=(
                b["bill_id"], b["patient_id"], b["appointment_id"], b["bill_date"],
                b["consultation_charge"], b["medicine_charge"], b["room_charge"], b["other_charge"],
                b["total_amount"], b["payment_status"], b["payment_method"]
            ))

    def refresh_medicines(self):
        for row in self.medicines_tree.get_children():
            self.medicines_tree.delete(row)
        for m in self.db.query("SELECT * FROM medicines ORDER BY medicine_id"):
            self.medicines_tree.insert("", "end", values=(
                m["medicine_id"], m["medicine_name"], m["category"], m["unit_price"],
                m["stock_quantity"], m["expiry_date"]
            ))

    def refresh_sales(self):
        for row in self.sales_tree.get_children():
            self.sales_tree.delete(row)
        rows = self.db.query("SELECT * FROM medicine_sales ORDER BY sale_date DESC")
        total = 0
        for s in rows:
            total += s["total_amount"] or 0
            self.sales_tree.insert("", "end", values=(
                s["sale_id"], s["medicine_id"], s["patient_id"], s["quantity"],
                s["sale_date"], s["total_amount"]
            ))
        self.pharmacy_revenue_lbl.config(text=f"Total Pharmacy Revenue: {total}")

    def refresh_patients(self):
        for row in self.patients_tree.get_children():
            self.patients_tree.delete(row)
        for p in self.db.query("SELECT * FROM patients ORDER BY patient_id"):
            self.patients_tree.insert("", "end", values=(
                p["patient_id"], p["first_name"], p["last_name"], p["date_of_birth"],
                p["gender"], p["phone"], p["email"], p["address"], p["blood_group"]
            ))

    def refresh_staff(self):
        for row in self.staff_tree.get_children():
            self.staff_tree.delete(row)
        for s in self.db.query("SELECT * FROM staff ORDER BY staff_id"):
            self.staff_tree.insert("", "end", values=(
                s["staff_id"], s["first_name"], s["last_name"], s["role"],
                s["department"], s["phone"], s["email"], s["hire_date"]
            ))

    def refresh_appointments(self):
        for row in self.appt_tree.get_children():
            self.appt_tree.delete(row)
        rows = self.db.query("""
            SELECT a.appointment_id,
                   p.first_name || ' ' || p.last_name AS patient,
                   s.first_name || ' ' || s.last_name AS staff,
                   a.appointment_date, a.appointment_time, a.reason, a.status
            FROM appointments a
            JOIN patients p ON a.patient_id = p.patient_id
            JOIN staff s ON a.staff_id = s.staff_id
            ORDER BY a.appointment_date DESC, a.appointment_time
        """)
        for a in rows:
            self.appt_tree.insert("", "end", values=(
                a["appointment_id"], a["patient"], a["staff"], a["appointment_date"],
                a["appointment_time"], a["reason"], a["status"]
            ))

    @staticmethod
    def clear_entries(entries_dict):
        for widget in entries_dict.values():
            if isinstance(widget, ttk.Combobox):
                widget.set("")
            else:
                widget.delete(0, "end")


if __name__ == "__main__":
    app = HospitalApp()
    app.mainloop()
